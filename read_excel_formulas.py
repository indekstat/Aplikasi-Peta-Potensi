import openpyxl

wb = openpyxl.load_workbook('Contoh-Perhitungan.xlsx', data_only=False)

for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    print(f"--- Sheet: {sheet_name} ---")
    for row in range(1, 10):
        row_vals = []
        for col in range(1, min(sheet.max_column + 1, 15)):
            cell = sheet.cell(row=row, column=col)
            if cell.data_type == 'f':
                row_vals.append(f"FORMULA: {cell.value}")
            else:
                row_vals.append(str(cell.value))
        print(" | ".join(row_vals))
    print("\n")
